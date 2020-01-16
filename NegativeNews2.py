import pandas as pd
import numpy as np
import math
import sys
import openpyxl

from openpyxl import Workbook
import win32com.client
from win32com.client import Dispatch


# wb = Workbook('data/股票负面^^^20191227.xlsx')
# negativenews = wb.worksheets[0]
#negative = openpyxl.load_workbook('data/negativenews20200106.xlsx')
score = openpyxl.load_workbook('data/newdata_stock20200115_standard.xlsx')
sheetssc = score.sheetnames
sc = score[sheetssc[1]]

sheetsne = score.sheetnames
negativenews = score[sheetsne[3]]

File = open("data/negativenews20200115.txt", "w", encoding=u'utf-8', errors='ignore')
File.write("标题"+"," + "风险类别" +","+ "重要性" +","+ "证券代码" +","+ "证券简称" +","+ "公司全称"+","+ "来源" +","+ "时间"+"\n")

i = 2
j = 2
for a in range(396):
    ne = negativenews.cell(row=i, column=6).value
    for b in range(200):
        ser = sc.cell(row=j, column=3).value
        j = j+1
        k = 1
        if ser==ne:
            File.write(str(negativenews.cell(row=i, column=2).value) + ","+ \
                       str(negativenews.cell(row=i, column=3).value) + "," + str(negativenews.cell(row=i, column=4).value)+ ","+ \
                       str(negativenews.cell(row=i, column=5).value) + "," + str(negativenews.cell(row=i, column=6).value)+ ","+ \
                       str(negativenews.cell(row=i, column=7).value) + "," + str(negativenews.cell(row=i, column=8).value) + "," + \
                       str(negativenews.cell(row=i, column=9).value) + "," + "\n")
    j = 2
    i = i + 1
print("ok")


