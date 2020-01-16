import pandas as pd
import numpy as np
import math
import sys
import openpyxl

from openpyxl import Workbook
import win32com.client
from win32com.client import Dispatch


# wb = Workbook('data/股票负面^^^20191227.xlsx')
# negativenews = wb.worksheets[0]
negative = openpyxl.load_workbook('data/newdata_stock20200115_standard.xlsx')
sheetsne = negative.sheetnames
negativeinds = negative[sheetsne[4]]


score = openpyxl.load_workbook('data/股票所属行业.xlsx')
sheetssc = score.sheetnames
sc = score[sheetssc[0]]

File = open("data/negativeindustry20200115.txt", "w", encoding=u'utf-8', errors='ignore')
File.write("股票名称"+"," + "所属行业" +"\n")

i = 2
j = 2
for a in range(200):
    ne = negativeinds.cell(row=i, column=4).value
    for b in range(3635):
        ser = sc.cell(row=j, column=3).value
        j = j+1
        k = 1
        if ser==ne:
            File.write(str(sc.cell(row=j, column=2).value) + ","+ str(sc.cell(row=j, column=3).value) + ","+"\n")
    j = 2
    i = i + 1
print("ok")


